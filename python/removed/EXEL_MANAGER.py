from openpyxl import load_workbook
TYPE_EXL = '.xlsx'
EXL_FOLDER_PATH = './PYTHON/EXCEL/'#엑셀 폴더 경로
CANT_FIND = '대상을 찾을 수 없습니다'

#엑셀 관리 클래스, openpyxl 라이브러리 사용, 개인 스타일링을 위해 클래스로 구현
class EXEL_MANAGER:
    def __init__(self):
        self.exl = None
        self.col_idx = tuple( ':',(chr(i) for i in range(65, 91) if True))
    
    #엑셀 읽어오기, 굳이 리턴 필요한가?, 시트는 미지정시 1번시트 반환
    def read_exl(self, exl_name, sheet_name=None, column=0, row=0):
        #경로 지정
        path = EXL_FOLDER_PATH+exl_name+TYPE_EXL
        try:
            if(self.exl!=None):
                self.exl.save(path)
            self.exl = load_workbook(path)
            ws = self.exl[sheet_name] if (sheet_name != None) else self.exl.active
            #셀 인덱싱 알고리즘
            col = self.col_idx[column%26]
            ws = ws[col+str(row+1)]
        except Exception as e:
            #만약에 경로, 시트를 못찾거나 IO 에러 발생시 에러처리
            print(f"Exception type: {type(e)}")
            print(f"Exception message: {str(e)}")
            raise e
        return self.get_data(ws)
        
    #엑셀의 데이터를 가져오는 함수
    def get_data(ws):
        #특정 시트를 . 다읽을수도 있고, 특정 셀만 읽을수도 있음(행=0, 열=0, 지정 가능, 미지정 속성은 전체 읽어오기)
        #읽어왔을때 값이 없으면 CANT_FIND 반환, 전체 읽을때 길이 지정이 중요할거 같은데...
        return ws.value if (ws.value != None) else CANT_FIND

    #엑셀에 데이터를 내보내는 함수
    def write_exl(self, exl_name, data=None, sheet_name=None, column=0, row=0):
        #경로 지정
        path = EXL_FOLDER_PATH+exl_name+TYPE_EXL
        #엑셀 파일이 없으면 생성 있으면 저장후 생성
        try:
            if(self.exl!=None):
                self.exl.save(path)
            self.exl = load_workbook(path)
            ws = self.exl[sheet_name] if (sheet_name != None) else self.exl.active
            col = self.col_idx[column%27]
            row = row if (row != 0) else self.col_idx[row]
            ws = ws[col+str(row+1)]
            ws.value = data
            self.exl.save(path)
        except Exception as e:
            #만약에 경로, 시트를 못찾거나 IO 에러 발생시 에러처리
            print(f"Exception type: {type(e)}")
            print(f"Exception message: {str(e)}")
            raise e
